"""Steerco: monthly snapshots, 12-month aggregation, backfill, lang-aware one-pager."""
from app.models import SteercoEntry
from app.routers.steerco import (I18N, _aggregate, month_keys, _onepager,
                                 _render_pptx, _svg_line_chart)
from tests.conftest import login


def _enable(client):
    """Steerco module is off by default; turn it on as admin."""
    login(client, "admin@test")
    r = client.put("/api/admin/modules-config", json={"steerco": {"enabled": True}})
    assert r.status_code == 200, r.text


# ---- pure helpers --------------------------------------------------------------

def test_month_keys():
    assert month_keys("2026-02", 3) == ["2025-12", "2026-01", "2026-02"]
    assert len(month_keys("2026-07", 12)) == 12


def test_svg_handles_gaps_and_empty():
    svg = _svg_line_chart({"labels": ["a", "b", "c"], "y_max": 100,
                           "series": [{"name": "x", "color": "#000", "data": [None, 40, 50]}]}, "no data")
    assert svg.startswith("<svg") and "<path" in svg and "<circle" not in svg   # line only, no markers
    assert "no data" in _svg_line_chart({"series": []}, "no data")


# ---- aggregation ---------------------------------------------------------------

def test_aggregate_builds_calendar_year_charts_and_annual_sla(db, seeded):
    sid = seeded["squad_a"]
    for p, users, inc in [("2026-05", 200, 10), ("2026-06", 210, 15), ("2026-07", 247, 13)]:
        db.add(SteercoEntry(squad_id=sid, period=p, data={
            "kpis": [{"label": "Users", "value": str(users), "trend": "up"}],
            "sla": {"services": ["Incidents"], "cells": [{"v": "99,4%", "s": "ok"}]},
            "incidents": str(inc)}))
    db.commit()
    rd = _aggregate(db, sid, "2026-07")
    # 12 columns = the calendar year, January to December: the chart always starts in
    # January (empty until May here), and the report month July sits at index 6.
    assert len(rd["kpi_chart"]["labels"]) == 12 and rd["kpi_chart"]["labels"][0] == "01/26"
    users_series = rd["kpi_chart"]["series"][0]
    assert users_series["data"][0] is None                       # January, no data yet
    assert users_series["data"][7] is None                       # August (future), empty
    seen = [v for v in users_series["data"] if v is not None]
    assert seen[0] == 100 and seen[-1] > 100                     # indexed to base 100, growing
    assert rd["incidents_chart"]["series"][0]["data"][6] == 13   # incidents at July, not the last column
    rows = rd["sla"]["rows"]
    assert [r["period"] for r in rows] == ["__current__", "__trailing__"]         # current + annual avg
    assert rd["kpis"][0]["value"] == "247"                                        # cards = current month


def test_change_vs_m1_and_sla_colour_are_computed(db, seeded):
    """The squad leader enters raw numbers only: the variation vs M-1 comes from the
    previous month's snapshot, the SLA colour from the value (>90 green, 80-90 amber,
    <80 red). Anything stored by hand is overridden."""
    sid = seeded["squad_a"]
    db.add(SteercoEntry(squad_id=sid, period="2026-06", data={
        "kpis": [{"label": "Users", "value": "235"}, {"label": "K8aaS", "value": "8"},
                 {"label": "DBaaS", "value": "4"}]}))
    db.add(SteercoEntry(squad_id=sid, period="2026-07", data={
        # stale hand-entered values on purpose: they must be recomputed
        "kpis": [{"label": "Users", "value": "247", "trend": "down", "delta": "-99"},
                 {"label": "K8aaS", "value": "5"}, {"label": "DBaaS", "value": "4"},
                 {"label": "New", "value": "3"}],
        "sla": {"services": ["A", "B", "C", "D", "E"],
                "cells": [{"v": "99,4%", "s": "ko"}, {"v": "90%"}, {"v": "84,5%"},
                          {"v": "72%"}, {"v": ""}]}}))
    db.commit()
    rd = _aggregate(db, sid, "2026-07")
    assert [(k["trend"], k["delta"]) for k in rd["kpis"]] == [
        ("up", "+12"),      # 247 vs 235
        ("down", "-3"),     # 8 -> 5
        ("flat", "0"),      # unchanged
        ("flat", ""),       # no previous value, no variation shown
    ]
    cur = rd["sla"]["rows"][0]["cells"]
    assert [c["s"] for c in cur] == ["ok", "warn", "warn", "ko", None]   # 90 stays amber
    html = _onepager("Squad A", "2026-07", rd, I18N["en"])
    assert "▲ +12" in html and "b-ko" in html and "b-warn" in html


def test_sla_percentages_are_capped_at_100(client, db, seeded):
    """A SLA is a percentage: 100 is the ceiling, whatever is sent (typo, import)."""
    _enable(client)
    sid = seeded["squad_a"]
    login(client, seeded["sl_a"])
    assert client.put(f"/api/steerco/squad/{sid}/enabled", json={"enabled": True}).status_code == 200
    r = client.put(f"/api/steerco/squad/{sid}?period=2026-07", json={
        "sla": {"services": ["A", "B"], "cells": [{"v": "994%"}, {"v": "99,4%"}]}})
    assert r.status_code == 200
    assert [c["v"] for c in r.json()["data"]["sla"]["cells"]] == ["100%", "99,4%"]
    rd = _aggregate(db, sid, "2026-07")
    assert [c["v"] for c in rd["sla"]["rows"][0]["cells"]] == ["100%", "99,4%"]


def test_onepager_layout_and_translation(db, seeded):
    sid = seeded["squad_a"]
    db.add(SteercoEntry(squad_id=sid, period="2026-07", data={
        "kpis": [{"label": "Users", "value": "247"}],
        "sla": {"services": ["Incidents"], "cells": [{"v": "99,4%", "s": "ok"}]},
        "incidents": "13"}))
    db.commit()
    rd = _aggregate(db, sid, "2026-07")
    html_en = _onepager("Squad A", "2026-07", rd, I18N["en"])
    # left column (KPIs) comes before the SLA table; right column has the KPI chart
    assert html_en.index("kpi-row") < html_en.index(">SLA")   # KPI panel before the SLA panel
    assert "hdr" in html_en and "July 2026" in html_en          # header + spelled-out month
    # SLA rows are the current month + the annual average; the chart sub shows the year.
    assert "Current month" in html_en and "Annual average" in html_en and "KPI trend" in html_en
    assert "2026, base 100" in html_en                          # chart sub-label carries the year
    html_fr = _onepager("Squad A", "2026-07", rd, I18N["fr"])
    assert "Mois en cours" in html_fr and "Moyenne annuelle" in html_fr and "Évolution KPI" in html_fr


# ---- API -----------------------------------------------------------------------

def test_module_gate_blocks_when_disabled(client, seeded):
    login(client, "admin@test")
    assert client.get("/api/steerco/entries?period=2026-07").status_code == 404


def test_snapshot_backfill_history_and_document(client, seeded):
    _enable(client)
    sid = seeded["squad_a"]
    login(client, seeded["sl_a"])
    assert client.put(f"/api/steerco/squad/{sid}/enabled", json={"enabled": True}).status_code == 200
    # this month's snapshot
    assert client.put(f"/api/steerco/squad/{sid}?period=2026-07", json={
        "kpis": [{"label": "Users", "value": "247"}],
        "sla": {"services": ["Incidents"], "cells": [{"v": "99,4%", "s": "ok"}]},
        "incidents": "13"}).status_code == 200
    # backfill two past months in one shot
    r = client.put(f"/api/steerco/squad/{sid}/history", json={"months": {
        "2026-06": {"kpis": [{"label": "Users", "value": "210"}], "incidents": "15"},
        "2026-05": {"kpis": [{"label": "Users", "value": "200"}], "incidents": "10"}}})
    assert r.status_code == 200 and r.json()["count"] == 2
    hist = client.get(f"/api/steerco/squad/{sid}/history?period=2026-07").json()
    assert len(hist["months"]) == 12

    # leadership one-pager, in both languages, plus the PPTX
    login(client, "admin@test")
    en = client.get(f"/api/steerco/onepager.html?squad_id={sid}&period=2026-07&lang=en")
    assert en.status_code == 200 and "KPI trend" in en.text and "247" in en.text
    fr = client.get(f"/api/steerco/onepager.html?squad_id={sid}&period=2026-07&lang=fr")
    assert "Évolution KPI" in fr.text
    assert client.get(f"/api/steerco/document.pptx?period=2026-07&lang=en").status_code == 200


def test_squad_detail_reflects_steerco_enabled(client, seeded):
    """Regression: SquadDetail (used by the My Squads cards + Reporting section to
    read the toggle) must carry steerco_enabled, else activation looks like a no-op."""
    _enable(client)
    sid = seeded["squad_a"]
    login(client, seeded["sl_a"])
    assert client.get(f"/api/squads/{sid}").json()["steerco_enabled"] is False
    assert client.put(f"/api/squads/{sid}", json={"steerco_enabled": True}).status_code == 200
    assert client.get(f"/api/squads/{sid}").json()["steerco_enabled"] is True   # persisted + serialized


def test_entry_reports_monthly_fill_status(client, seeded):
    """The reporting launcher relies on filled/updated_at/updated_by to show whether
    this month's Steerco is done or still to do."""
    _enable(client)
    sid = seeded["squad_a"]
    login(client, seeded["sl_a"])
    before = client.get(f"/api/steerco/squad/{sid}?period=2026-07").json()
    assert before["filled"] is False and before["updated_at"] is None
    client.put(f"/api/steerco/squad/{sid}?period=2026-07", json={"kpis": [{"label": "Cloud Users", "value": "1"}]})
    after = client.get(f"/api/steerco/squad/{sid}?period=2026-07").json()
    assert after["filled"] is True and after["updated_at"] and after["updated_by"] == "SL A"


def test_preview_uses_unsaved_data_without_persisting(client, seeded):
    """The wizard preview renders the still-unsaved snapshot (squad-leader accessible)
    and must not persist it."""
    _enable(client)
    sid = seeded["squad_a"]
    login(client, seeded["sl_a"])
    assert client.put(f"/api/steerco/squad/{sid}/enabled", json={"enabled": True}).status_code == 200
    r = client.post(f"/api/steerco/squad/{sid}/preview.html?period=2026-07&lang=en",
                    json={"kpis": [{"label": "Users", "value": "997"}],
                          "sla": {"services": ["Incidents"], "cells": [{"v": "99,4%", "s": "ok"}]},
                          "incidents": "3"})
    assert r.status_code == 200 and "997" in r.text          # unsaved value shows in the preview
    # nothing was persisted for that period
    assert client.get(f"/api/steerco/squad/{sid}?period=2026-07").json()["data"] in (None, {})


def test_template_columns_are_the_charted_window(seeded):
    """The workbook's 12 month columns must be exactly the window the one-pager charts
    read: the report year's calendar months, January to December, so the filled months
    line up with the charted months and the charts start in January."""
    import io
    from openpyxl import load_workbook
    from app.steerco_import import _month_label, template_bytes
    from app.routers.steerco import year_months

    wb = load_workbook(io.BytesIO(template_bytes("2026-07")))
    headers = [wb["KPIs"].cell(1, 2 + j).value for j in range(12)]
    expected = [_month_label(k) for k in year_months("2026-07")]
    expected[6] += "*"                                  # July is the report month
    assert headers == expected
    assert headers[0] == "Janv 26"                      # always January, not a rolling start
    assert headers[6].endswith("*") and headers[-1] == "Déc 26"


def test_excel_template_and_import(client, seeded, db):
    """Admin downloads the blank template, uploads a filled one, and it lands as
    SteercoEntry snapshots for the named squad (Admin > Import)."""
    import io
    from openpyxl import load_workbook
    from app.steerco_import import template_bytes

    login(client, "admin@test")
    # template download
    r = client.get("/api/admin/import-steerco/template")
    assert r.status_code == 200 and r.content[:2] == b"PK"        # a real .xlsx (zip)

    # fill it: squad name + a couple of months of one KPI + current SLA + an event.
    # Columns are the calendar year: B = Jan ... G = Jun, H = Jul (the report month).
    wb = load_workbook(io.BytesIO(template_bytes("2026-07")))
    wb["Infos"]["B2"] = "Squad A"; wb["Infos"]["B3"] = "2026-07"
    wb["KPIs"]["G2"] = 118; wb["KPIs"]["H2"] = 120          # Cloud Users Jun, Jul (current)
    wb["SLA"]["H2"] = 99.2                                   # Incidents, current month
    wb["Incidents"]["H2"] = 4
    wb["Evenements passes"]["A2"] = "12/07"; wb["Evenements passes"]["C2"] = "Incident majeur"
    wb["Evenements passes"]["D2"] = "Attention"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    r = client.post("/api/admin/import-steerco",
                    files={"file": ("steerco.xlsx", buf.getvalue(),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    s = r.json()
    assert s["squad"] == "Squad A" and s["period"] == "2026-07" and s["months"] >= 2

    # persisted + feeds the one-pager
    from app.routers.steerco import _aggregate
    rd = _aggregate(db, seeded["squad_a"], "2026-07")
    assert rd["kpis"][0]["value"] == "120"
    assert rd["sla"]["rows"][0]["cells"][0]["s"] == "ok"       # 99,2% -> green, computed
    seen = [v for v in rd["kpi_chart"]["series"][0]["data"] if v is not None]
    assert seen[0] == 100 and len(seen) >= 2                # Jun + Jul indexed
    # The imported event's severity reaches the one-pager as a coloured chip.
    html = _onepager("Squad A", "2026-07", rd, I18N["fr"])
    assert "Incident majeur" in html and "#FBF0D9" in html   # "Attention" -> amber chip


def test_template_matches_what_the_squad_actually_reports(client, db, seeded):
    """A squad's workbook must propose ITS rows (name pre-filled, its KPI / SLA lines),
    not a canned list, otherwise the file and the app describe different squads."""
    import io
    from openpyxl import load_workbook

    sid = seeded["squad_a"]
    db.add(SteercoEntry(squad_id=sid, period="2026-07", data={
        "kpis": [{"label": "Cloud Users", "value": "247"}, {"label": "Terraform", "value": "42"}],
        "sla": {"services": ["Incidents", "Vault"], "cells": [{"v": "99%"}, {"v": "95%"}]}}))
    db.commit()

    login(client, "admin@test")
    r = client.get(f"/api/admin/import-steerco/template?squad_id={sid}")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert wb["Infos"]["B2"].value == "Squad A"                       # name pre-filled
    assert [wb["KPIs"].cell(i, 1).value for i in (2, 3)] == ["Cloud Users", "Terraform"]
    assert [wb["SLA"].cell(i, 1).value for i in (2, 3)] == ["Incidents", "Vault"]
    # without a squad, the standard structure is still served
    generic = load_workbook(io.BytesIO(client.get("/api/admin/import-steerco/template").content))
    assert generic["KPIs"]["A2"].value == "Cloud Users" and not generic["Infos"]["B2"].value


def test_import_never_deletes_what_was_entered_in_the_app(client, db, seeded):
    """The workbook adds and updates; it must not wipe a KPI, an SLA service or the
    events a squad entered in the app just because the file does not mention them."""
    import io
    from openpyxl import load_workbook
    from app.steerco_import import template_bytes

    sid = seeded["squad_a"]
    db.add(SteercoEntry(squad_id=sid, period="2026-07", data={
        "kpis": [{"label": "Cloud Users", "value": "247"}, {"label": "Terraform", "value": "42"}],
        "sla": {"services": ["Vault"], "cells": [{"v": "95,0%"}]},
        "last_events": [{"date": "01/07", "text": "Saisi dans l'app"}]}))
    db.commit()

    # the generic template knows nothing about Terraform, Vault or the event.
    # Calendar-year columns: July (the report month) is column 8 (H).
    wb = load_workbook(io.BytesIO(template_bytes("2026-07")))
    wb["Infos"]["B2"] = "Squad A"; wb["Infos"]["B3"] = "2026-07"
    wb["KPIs"].cell(2, 8, 250)                       # updates Cloud Users only (July column)
    wb["KPIs"]["A7"] = "Nouveau KPI"; wb["KPIs"].cell(7, 8, 5)      # a row added by hand
    buf = io.BytesIO(); wb.save(buf)

    login(client, "admin@test")
    r = client.post("/api/admin/import-steerco",
                    files={"file": ("s.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 200, r.text
    assert r.json()["kept_kpis"] == ["Terraform"]     # preserved AND reported to the admin

    db.expire_all()
    data = (db.query(SteercoEntry)
            .filter(SteercoEntry.squad_id == sid, SteercoEntry.period == "2026-07").one()).data
    by_label = {k["label"]: k["value"] for k in data["kpis"]}
    assert by_label["Cloud Users"] == "250"          # updated by the file
    assert by_label["Terraform"] == "42"             # untouched, not deleted
    assert by_label["Nouveau KPI"] == "5"            # a row added in Excel is picked up
    assert data["sla"]["services"] == ["Vault"]      # service kept, footnote not read as one
    assert len(data["last_events"]) == 1             # empty event sheet does not wipe events


def test_import_rejects_a_malformed_report_month(seeded):
    """The report month drives the 12 collected columns, so a non "AAAA-MM" value must
    fail loudly instead of silently importing into the wrong months."""
    import io
    import pytest
    from openpyxl import load_workbook
    from app.steerco_import import parse_workbook, template_bytes

    wb = load_workbook(io.BytesIO(template_bytes("2026-07")))
    wb["Infos"]["B2"] = "Squad A"; wb["Infos"]["B3"] = "juillet 2026"
    buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(ValueError, match="AAAA-MM"):
        parse_workbook(buf.getvalue())


def test_import_unknown_squad_is_rejected(client, seeded):
    import io
    from openpyxl import load_workbook
    from app.steerco_import import template_bytes

    login(client, "admin@test")
    wb = load_workbook(io.BytesIO(template_bytes()))
    wb["Infos"]["B2"] = "No Such Squad"; wb["Infos"]["B3"] = "2026-07"
    buf = io.BytesIO(); wb.save(buf)
    r = client.post("/api/admin/import-steerco",
                    files={"file": ("s.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 400 and "introuvable" in r.json()["detail"].lower()


def test_backfill_requires_edit_rights(client, seeded):
    _enable(client)
    login(client, seeded["member"])
    assert client.put(f"/api/steerco/squad/{seeded['squad_a']}/history",
                      json={"months": {"2026-06": {"incidents": "3"}}}).status_code in (403, 404)
