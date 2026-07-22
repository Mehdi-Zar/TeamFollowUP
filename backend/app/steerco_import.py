"""Steerco data collection via Excel (Admin > Import), mirroring app/import_org.py.

An admin downloads a blank workbook (:func:`template_bytes`), a squad fills it in,
and the admin uploads it back (:func:`import_steerco`). The file is parsed in memory
and written straight to ``SteercoEntry`` rows (12 monthly snapshots for the charts +
the full current-month snapshot with its events). No image rebuild, idempotent:
re-uploading replaces that squad's monthly data.

Only raw numbers are collected: the KPI variation vs M-1 and the SLA colours are
computed from the values themselves when the one-pager is rendered (see
``routers/steerco.py``), so the workbook has no trend / variation / status column.

The workbook has 7 sheets: Instructions, Infos, KPIs, SLA, Incidents,
Evenements passes, Evenements a venir. Keep the sheet names / header rows stable.

The 12 monthly columns are the ROLLING 12 months ending at the report month (the
report month is the last one, marked with a *), not a calendar year: that is exactly
the window the one-pager charts read (``routers.steerco.month_keys``), so a filled
workbook produces complete charts instead of leading gaps.
"""
import io
from datetime import date

from sqlalchemy.orm import Session

from .models import SteercoEntry, Squad
# Single source of truth for the 12-month window, shared with the renderer so the
# collected columns and the charted months can never drift apart.
from .routers.steerco import month_keys

# ---- structure -------------------------------------------------------------
MONTHS_ABBR_FR = ["Janv", "Févr", "Mars", "Avr", "Mai", "Juin",
                  "Juil", "Août", "Sept", "Oct", "Nov", "Déc"]
KPI_LABELS = ["Cloud Users", "Landing Zone", "K8aaS", "DBaaS", "Software Factory"]
SWF_SUBS = ["GitLab", "Artifactory", "SonarQube"]
_SWF_KEYS = {s.lower() for s in SWF_SUBS}      # how the "Infos" rows are matched back
SLA_SERVICES = ["Incidents", "Gitlab", "Artifactory", "Sonarqube"]
SEV_CHOICES = ["Critique", "Attention", "OK", "Prévu"]
SEV_MAP = {"critique": "red", "attention": "amber", "ok": "green", "prévu": "ice", "prevu": "ice"}


def default_period() -> str:
    """The month a blank template is built for: the current one ("YYYY-MM")."""
    today = date.today()
    return f"{today.year:04d}-{today.month:02d}"


def _month_label(key: str) -> str:
    """"2025-08" -> "Août 25" (column header of the workbook's month grid)."""
    y, m = key.split("-")[:2]
    return f"{MONTHS_ABBR_FR[int(m) - 1]} {y[2:]}"


# ---- value helpers ---------------------------------------------------------
def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("%", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _count(v):
    n = _num(v)
    return None if n is None else str(int(round(n)))


def _pct(v):
    """A SLA percentage, capped at 100 (and never negative)."""
    n = _num(v)
    return None if n is None else (f"{min(max(n, 0.0), 100.0):.1f}".replace(".", ",") + "%")


# ============================================================================
# Template generation
# ============================================================================
def structure_for_squad(db: Session, squad_id: int) -> tuple[list[str], list[str]]:
    """The KPI labels and SLA services a squad currently reports, read from its most
    recent snapshot, so the workbook proposes *its* rows and not a canned list.
    Falls back to the standard structure when the squad has never reported."""
    rows = (db.query(SteercoEntry)
            .filter(SteercoEntry.squad_id == squad_id)
            .order_by(SteercoEntry.period.desc()).all())
    kpis = next(([k.get("label") for k in (r.data or {}).get("kpis") or [] if k.get("label")]
                 for r in rows if (r.data or {}).get("kpis")), None)
    services = next(((((r.data or {}).get("sla") or {}).get("services") or [])
                     for r in rows if ((r.data or {}).get("sla") or {}).get("services")), None)
    return (kpis or list(KPI_LABELS)), (services or list(SLA_SERVICES))


def build_template_workbook(period: str | None = None, kpi_labels: list[str] | None = None,
                            sla_services: list[str] | None = None, squad_name: str = ""):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    NAVY, ICE = "1E2761", "E8F0FE"
    HEAD = Font(bold=True, color="FFFFFF", size=11)
    HEAD_FILL = PatternFill("solid", fgColor=NAVY)
    LBL = Font(bold=True, color="1E2761")
    MUTED = Font(color="64748B", italic=True, size=10)
    CUR_FILL = PatternFill("solid", fgColor=ICE)
    THIN = Side(style="thin", color="D9E1EC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")
    period = period or default_period()
    kpi_labels = list(kpi_labels or KPI_LABELS)
    sla_services = list(sla_services or SLA_SERVICES)
    # Rolling window: the report month is the LAST of the 12 columns.
    keys = month_keys(period, 12)
    cur_idx = len(keys)

    def hrow(ws, row, values, widths=None):
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = HEAD; c.fill = HEAD_FILL; c.alignment = CENTER; c.border = BORDER
        if widths:
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[row].height = 22

    def month_headers(ws, start_col):
        for j, key in enumerate(keys):
            col = start_col + j
            m = _month_label(key)
            c = ws.cell(row=1, column=col, value=(f"{m}*" if j + 1 == cur_idx else m))
            c.font = HEAD; c.fill = HEAD_FILL; c.alignment = CENTER; c.border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = 11

    def fill_cells(ws, row, start_col, n):
        for j in range(n):
            c = ws.cell(row=row, column=start_col + j)
            c.alignment = CENTER; c.border = BORDER
            if j + 1 == cur_idx:
                c.fill = CUR_FILL

    wb = Workbook()

    ws = wb.active; ws.title = "Instructions"; ws.column_dimensions["A"].width = 110
    lines = [
        ("Steerco - fichier de collecte des données", LBL), ("", None),
        ("Rassemblez ici toutes les données d'un rapport Steerco mensuel (KPI, SLA, incidents, "
         "évènements). Une fois rempli, réimportez le fichier dans l'app (Admin > Import).", None), ("", None),
        ("Comment remplir :", LBL),
        ("1) Onglet « Infos » : nom exact de la squad, mois du rapport en cours (AAAA-MM), "
         "et les 3 sous-métriques Software Factory du mois en cours.", None),
        ("2) Onglet « KPIs » : une valeur par mois (le mois en cours est le dernier, marqué d'une *). "
         "La variation vs M-1 est calculée automatiquement, rien à saisir.", None),
        ("3) Onglet « SLA » : le % par mois (100 % maximum). La couleur est calculée "
         "automatiquement (au-dessus de 90 % vert, de 80 à 90 % orange, en dessous de 80 % rouge).", None),
        ("4) Onglet « Incidents » : le nombre d'incidents par mois.", None),
        ("5) Onglets « Evenements passes / a venir » : Date, Type, Libellé, Gravité.", None), ("", None),
        ("Valeurs autorisées :", LBL),
        ("- Gravité : Critique / Attention / OK / Prévu", None), ("", None),
        ("Les colonnes couvrent les 12 mois glissants qui se terminent au mois du rapport, "
         "soit exactement la fenêtre des graphiques du one-pager. Les mois vides sont ignorés.", MUTED),
        ("Si vous changez le mois du rapport, retéléchargez le modèle : les colonnes se décalent.", MUTED),
        ("Ne renommez pas les onglets ni les en-têtes : l'import s'appuie dessus.", MUTED),
    ]
    for r, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=r, column=1, value=text); c.alignment = Alignment(wrap_text=True, vertical="top")
        if font:
            c.font = font

    ws = wb.create_sheet("Infos")
    ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 26
    hrow(ws, 1, ["Champ", "Valeur"])
    infos = [("Squad (nom exact dans l'app)", squad_name),
             ("Mois du rapport en cours (AAAA-MM)", period),
             ("", ""),
             ("Software Factory - sous-métriques (mois en cours)", ""),
             *[(s, "") for s in SWF_SUBS]]
    for i, (k, v) in enumerate(infos, start=2):
        a = ws.cell(row=i, column=1, value=k); a.alignment = LEFT
        if k:
            a.border = BORDER
            if not v and "Software Factory" in k:
                a.font = LBL
        b = ws.cell(row=i, column=2, value=v); b.alignment = LEFT
        if k:
            b.border = BORDER

    ws = wb.create_sheet("KPIs"); ws.column_dimensions["A"].width = 20
    hrow(ws, 1, ["KPI"]); month_headers(ws, 2)
    for i, label in enumerate(kpi_labels, start=2):
        c = ws.cell(row=i, column=1, value=label); c.font = LBL; c.alignment = LEFT; c.border = BORDER
        fill_cells(ws, i, 2, 12)
    ws.cell(row=len(kpi_labels) + 3, column=1,
            value=("* = mois du rapport en cours (dernière colonne). La variation vs M-1 "
                   "est calculée automatiquement.")).font = MUTED

    ws = wb.create_sheet("SLA"); ws.column_dimensions["A"].width = 20
    hrow(ws, 1, ["Service"]); month_headers(ws, 2)
    for i, svc in enumerate(sla_services, start=2):
        c = ws.cell(row=i, column=1, value=svc); c.font = LBL; c.alignment = LEFT; c.border = BORDER
        fill_cells(ws, i, 2, 12)
    # Excel refuses a SLA above 100% straight in the file (the import caps it too).
    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=100, allow_blank=True)
    dv.errorTitle = "Pourcentage invalide"
    dv.error = "Une valeur SLA est un pourcentage entre 0 et 100."
    ws.add_data_validation(dv); dv.add(f"B2:M{1 + len(sla_services)}")
    # Note rows are prefixed with "*": that is what stops the row reader, so adding a
    # service row cannot make the footnote itself be read as a service name.
    ws.cell(row=len(sla_services) + 3, column=1,
            value=("* Valeurs en % (ex : 99,4), 100 % maximum. Couleur auto : au-dessus de 90 % vert, "
                   "de 80 à 90 % orange, en dessous de 80 % rouge. "
                   "Ajoutez une ligne pour suivre un service de plus.")).font = MUTED

    ws = wb.create_sheet("Incidents"); ws.column_dimensions["A"].width = 20
    hrow(ws, 1, [""]); month_headers(ws, 2)
    c = ws.cell(row=2, column=1, value="Nombre d'incidents"); c.font = LBL; c.alignment = LEFT; c.border = BORDER
    fill_cells(ws, 2, 2, 12)

    for title in ("Evenements passes", "Evenements a venir"):
        ws = wb.create_sheet(title)
        hrow(ws, 1, ["Date", "Type", "Libellé", "Gravité"], widths=[12, 20, 48, 16])
        for i in range(2, 8):
            for col in range(1, 5):
                ws.cell(row=i, column=col).border = BORDER
                ws.cell(row=i, column=col).alignment = LEFT if col == 3 else CENTER
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(SEV_CHOICES), allow_blank=True)
        ws.add_data_validation(dv); dv.add("D2:D7")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    return wb


def template_bytes(period: str | None = None, kpi_labels: list[str] | None = None,
                   sla_services: list[str] | None = None, squad_name: str = "") -> bytes:
    buf = io.BytesIO()
    build_template_workbook(period, kpi_labels, sla_services, squad_name).save(buf)
    return buf.getvalue()


# ============================================================================
# Parsing + import
# ============================================================================
def _rows_until_blank(ws, start=2, key_col=1):
    for r in range(start, ws.max_row + 1):
        k = ws.cell(r, key_col).value
        if k is None or not str(k).strip() or str(k).strip().startswith("*"):
            break
        yield r, str(k).strip()


def parse_workbook(content: bytes) -> dict:
    """Parse a filled workbook into {squad, period, months, counts}.

    ``months`` maps each "YYYY-MM" with data to its snapshot; the current period's
    entry is the full snapshot (KPI values + sub-metrics, SLA values, events)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True)
    required = {"Infos", "KPIs", "SLA", "Incidents", "Evenements passes", "Evenements a venir"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Onglets manquants dans le fichier : {', '.join(sorted(missing))}.")

    # Infos
    squad = period = None
    subs = {}
    ws = wb["Infos"]
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if not k:
            continue
        kl = str(k).strip().lower()
        if kl.startswith("squad"):
            squad = str(v).strip() if v else None
        elif "mois du rapport" in kl:
            period = str(v).strip() if v else None
        elif kl in _SWF_KEYS and v is not None and str(v).strip():
            subs[str(k).strip()] = _count(v)
    if not squad:
        raise ValueError("Renseignez le nom de la squad (onglet 'Infos').")
    if not period:
        raise ValueError("Renseignez le mois du rapport en cours (onglet 'Infos', format AAAA-MM).")
    try:
        y, m = (int(x) for x in period.split("-")[:2])
        if not (1 <= m <= 12):
            raise ValueError
        period = f"{y:04d}-{m:02d}"
    except ValueError:
        raise ValueError(f"Mois du rapport invalide : « {period} ». Format attendu : AAAA-MM.")
    # Same rolling window as the template and as the one-pager charts: the report
    # month is the last of the 12 columns.
    keys = month_keys(period, 12)
    cur_col = 2 + (len(keys) - 1)
    months = {k: {} for k in keys}

    # KPIs
    ws = wb["KPIs"]
    current_kpis = []
    for r, label in _rows_until_blank(ws):
        cur_val = None
        for j, key in enumerate(keys):
            cs = _count(ws.cell(r, 2 + j).value)
            if cs is not None:
                months[key].setdefault("kpis", []).append({"label": label, "value": cs})
                if 2 + j == cur_col:
                    cur_val = cs
        kpi = {"label": label, "value": cur_val or ""}
        if label.lower().startswith("software factory") and subs:
            kpi["sub"] = [{"label": k, "value": v} for k, v in subs.items()]
        current_kpis.append(kpi)

    # SLA
    ws = wb["SLA"]
    services = []
    sla_by_month = {k: {} for k in keys}
    for r, svc in _rows_until_blank(ws):
        services.append(svc)
        for j, key in enumerate(keys):
            n = _num(ws.cell(r, 2 + j).value)
            if n is not None:
                sla_by_month[key][svc] = {"v": _pct(n)}
    for key in keys:
        if sla_by_month[key]:
            months[key]["sla"] = {"services": services,
                                  "cells": [sla_by_month[key].get(s, {"v": ""}) for s in services]}
    current_sla = months.get(period, {}).get("sla") or {
        "services": services, "cells": [{"v": ""} for _ in services]}

    # Incidents
    ws = wb["Incidents"]
    for j, key in enumerate(keys):
        cs = _count(ws.cell(2, 2 + j).value)
        if cs is not None:
            months[key]["incidents"] = cs
    current_incidents = months.get(period, {}).get("incidents", "")

    # Events
    def read_events(sheet, default_sev):
        out = []
        w = wb[sheet]
        for r in range(2, w.max_row + 1):
            date = w.cell(r, 1).value
            text = w.cell(r, 3).value
            if not (date or text):
                continue
            out.append({
                "date": str(date).strip() if date else "",
                "tag": str(w.cell(r, 2).value).strip() if w.cell(r, 2).value else "",
                "text": str(text).strip() if text else "",
                "sev": SEV_MAP.get(str(w.cell(r, 4).value or "").strip().lower(), default_sev),
            })
        return out

    last_events = read_events("Evenements passes", "amber")
    next_events = read_events("Evenements a venir", "ice")

    # The current period gets the full snapshot (overrides its basic history entry).
    months[period] = {
        "kpis": current_kpis,
        "sla": current_sla,
        "incidents": current_incidents,
        "last_events": last_events,
        "next_events": next_events,
    }
    months = {k: v for k, v in months.items() if v}
    return {
        "squad": squad, "period": period, "months": months,
        "counts": {"months": len(months), "kpis": len(current_kpis), "services": len(services),
                   "events": len(last_events) + len(next_events)},
    }


# ---- merge rules -----------------------------------------------------------
# The import ADDS and UPDATES, it never deletes. A KPI or an SLA service entered in
# the app but absent from the workbook is kept, and a cell left empty in the file
# leaves the stored value alone. Removing something stays a one-click action in the
# app, which avoids the workbook silently wiping data it simply did not know about.

def _filled(v) -> bool:
    return v is not None and str(v).strip() != ""


def _merge_kpis(old: list, new: list) -> list:
    """Upsert the workbook's KPI values onto the stored list, matching labels
    case-insensitively. Unknown labels are appended, but only when they carry a
    value (an empty template row must not create a blank KPI)."""
    out = [dict(k) for k in (old or [])]
    index = {(k.get("label") or "").strip().lower(): i for i, k in enumerate(out)}
    for k in (new or []):
        key = (k.get("label") or "").strip().lower()
        if not key:
            continue
        sub = [s for s in (k.get("sub") or []) if _filled(s.get("value"))]
        if key in index:
            cur = out[index[key]]
            if _filled(k.get("value")):
                cur["value"] = k["value"]
            if sub:
                cur["sub"] = k["sub"]
        elif _filled(k.get("value")) or sub:
            out.append(dict(k))
            index[key] = len(out) - 1
    return out


def _merge_sla(old: dict, new: dict) -> dict:
    """Upsert the workbook's SLA values by service name, appending services the app
    did not know and keeping the ones the workbook did not mention."""
    services = list((old or {}).get("services") or [])
    cells = [dict(c) if isinstance(c, dict) else {"v": ""} for c in ((old or {}).get("cells") or [])]
    cells += [{"v": ""}] * (len(services) - len(cells))
    index = {s.strip().lower(): i for i, s in enumerate(services)}
    new_cells = (new or {}).get("cells") or []
    for j, svc in enumerate((new or {}).get("services") or []):
        value = (new_cells[j] or {}).get("v") if j < len(new_cells) else None
        key = (svc or "").strip().lower()
        if not key:
            continue
        if key in index:
            if _filled(value):
                cells[index[key]] = {"v": value}
        elif _filled(value):
            services.append(svc)
            cells.append({"v": value})
            index[key] = len(services) - 1
    return {"services": services, "cells": cells}


def _merge_month(old: dict, new: dict) -> dict:
    """One month's stored snapshot updated with what the workbook actually filled in."""
    out = dict(old or {})
    kpis = _merge_kpis(out.get("kpis") or [], new.get("kpis") or [])
    if kpis:
        out["kpis"] = kpis
    sla = _merge_sla(out.get("sla") or {}, new.get("sla") or {})
    if sla["services"]:
        out["sla"] = sla
    if _filled(new.get("incidents")):
        out["incidents"] = new["incidents"]
    # Events: an empty sheet means "not filled in", never "delete what is there".
    for key in ("last_events", "next_events"):
        if new.get(key):
            out[key] = new[key]
    return out


def import_steerco(db: Session, content: bytes, user_id: int | None = None) -> dict:
    """Parse a filled workbook and merge it into the squad's Steerco entries.

    Idempotent per (squad, period) and **non-destructive**: values present in the
    file are written, everything else the squad already reported is preserved (see
    the merge rules above). Enables Steerco on the squad. Returns a summary dict."""
    parsed = parse_workbook(content)
    name = parsed["squad"]
    # Case-insensitive exact match. `%`/`_` in a name would be LIKE wildcards, so the
    # comparison is done on the lowered value rather than with ilike().
    matches = [s for s in db.query(Squad).all() if s.name.strip().lower() == name.lower()]
    if not matches:
        raise ValueError(f"Squad introuvable : « {name} ». Vérifiez le nom exact dans l'app.")
    if len(matches) > 1:
        raise ValueError(f"Plusieurs squads s'appellent « {name} ». Renommez-en une pour lever l'ambiguïté.")
    squad = matches[0]
    squad.steerco_enabled = True
    kept = set()          # labels the app had and the workbook did not mention
    for period, data in parsed["months"].items():
        entry = (db.query(SteercoEntry)
                 .filter(SteercoEntry.squad_id == squad.id, SteercoEntry.period == period)
                 .one_or_none())
        if entry is None:
            db.add(SteercoEntry(squad_id=squad.id, period=period, data=_merge_month({}, data),
                                updated_by_user_id=user_id))
            continue
        in_file = {(k.get("label") or "").strip().lower() for k in (data.get("kpis") or [])}
        kept |= {k.get("label") for k in ((entry.data or {}).get("kpis") or [])
                 if k.get("label") and (k["label"]).strip().lower() not in in_file}
        entry.data = _merge_month(entry.data or {}, data)
        entry.updated_by_user_id = user_id
    db.flush()
    return {"squad": squad.name, "squad_id": squad.id, "period": parsed["period"],
            # KPIs the app already carried that the workbook did not cover: preserved,
            # and reported back so the admin sees the file was not the whole picture.
            "kept_kpis": sorted(kept), **parsed["counts"]}
