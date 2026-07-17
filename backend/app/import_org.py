"""Import a real organisation (tribe, squads, leaders, initiatives, OTDs) from an
Excel (or YAML) definition into the database.

Idempotent by natural key: an existing tribe (by name), user (by email), squad
(by name within the tribe) or initiative/OTD (by title within the tribe+year) is
UPDATED rather than duplicated, so it is safe to re-run and can push updates to an
already-populated environment. Nothing outside the described org is touched.

Primary path is the admin **upload API** (``POST /api/admin/import-org``, see
``routers/admin.py``): a filled Excel file is uploaded to the running app, parsed
in memory, and imported. No image rebuild and no redeploy are needed, and the same
flow works locally and in production (S3NS). ``download_org_template`` / the CLI
``--template`` produce the blank workbook to fill.

The importer core is reusable: ``read_upload(filename, bytes)`` parses an uploaded
Excel/YAML payload, and ``import_org(db, data)`` performs the idempotent import and
returns a summary. A CLI entry point (``main``) remains for offline imports from a
file already present in the container.

Leader accounts are created active and SSO-friendly: a later IdP login is matched
by email and inherits this account (its role and squad assignments), so pre-seeding
leaders here and wiring SSO afterwards compose cleanly.
"""
from __future__ import annotations

import logging
import os
import sys
from datetime import date, datetime, timezone

import yaml
from sqlalchemy import select
from sqlalchemy.orm import Session

from .database import SessionLocal
from .deps import SQUAD, TRIBE
from .models import Initiative, Otd, Squad, Tribe, User
from .security import hash_password

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
log = logging.getLogger("trt.import")

# Data files live in backend/data (copied to /app/data in the image). The input
# may be an Excel workbook (org.xlsx) or a YAML file (org.yaml); the default picks
# whichever exists, preferring Excel.
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")


def _default_path() -> str:
    """First existing default input (org.xlsx, then org.yaml/.yml); else org.xlsx."""
    for name in ("org.xlsx", "org.yaml", "org.yml"):
        p = os.path.join(DATA_DIR, name)
        if os.path.isfile(p):
            return p
    return os.path.join(DATA_DIR, "org.xlsx")

# Password set on leader accounts we create. It is only used for LOCAL password
# login; on an SSO environment the person is matched by email at first IdP login
# and never needs it. Override with IMPORT_DEFAULT_PASSWORD if you want a specific one.
DEFAULT_PASSWORD = os.environ.get("IMPORT_DEFAULT_PASSWORD", "changeme")


def _to_datetime(value) -> datetime | None:
    """Normalize a YAML date/datetime (or ISO string) to a tz-aware UTC datetime.

    PyYAML parses ``2026-12-31`` as a ``date`` and ``2026-12-31 09:00`` as a
    ``datetime``; strings are also accepted. Returns None for empty values.
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=timezone.utc)
    dt = datetime.fromisoformat(str(value))
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


def _get_or_create_user(db: Session, email: str | None, name: str | None,
                        role: str, tribe_id: int) -> tuple[User | None, bool]:
    """Return ``(user, created)`` for ``email``: create it (active, given
    role/tribe) if absent, otherwise sync its name/role/tribe. ``user`` is None
    when no email is provided (a leader is optional on a squad); ``created`` is
    True only when a new account was inserted."""
    email = (email or "").strip().lower()
    if not email:
        return None, False
    u = db.scalar(select(User).where(User.email == email))
    if u is None:
        u = User(email=email, display_name=(name or email), role=role,
                 status="active", tribe_id=tribe_id,
                 password_hash=hash_password(DEFAULT_PASSWORD))
        db.add(u)
        db.flush()
        log.info("User created: %s (%s)", email, role)
        return u, True
    # Keep the existing account but align it with the file.
    if name:
        u.display_name = name
    u.role = role
    u.tribe_id = tribe_id
    return u, False


def import_org(db: Session, data: dict) -> dict:
    """Create/update the whole organisation described by ``data`` (parsed from the
    Excel or YAML input). Runs in one transaction, commits at the end, and returns
    a summary of what was processed (for the API response / CLI log)."""
    year = int(data.get("year") or datetime.now(timezone.utc).year)
    created = {"users": 0, "squads": 0, "initiatives": 0, "otds": 0}

    # --- Tribe (matched by name) + its tribe leader ---
    tribe_def = data["tribe"]
    tribe = db.scalar(select(Tribe).where(Tribe.name == tribe_def["name"]))
    if tribe is None:
        tribe = Tribe(name=tribe_def["name"], description=tribe_def.get("description"), display_order=1)
        db.add(tribe)
        db.flush()
        log.info("Tribe created: %s", tribe.name)
    else:
        tribe.description = tribe_def.get("description", tribe.description)
    if tribe_def.get("leader"):
        _, c = _get_or_create_user(db, tribe_def["leader"].get("email"),
                                   tribe_def["leader"].get("name"), TRIBE, tribe.id)
        created["users"] += c

    # --- Squads (matched by name within the tribe) + their squad leaders ---
    squads_by_name: dict[str, Squad] = {}
    for i, sq in enumerate(data.get("squads", []) or [], start=1):
        leader, c = _get_or_create_user(db, (sq.get("leader") or {}).get("email"),
                                        (sq.get("leader") or {}).get("name"), SQUAD, tribe.id)
        created["users"] += c
        fields = dict(
            name=sq["name"], tribe_id=tribe.id, description=sq.get("description"),
            leader_user_id=leader.id if leader else None,
            display_order=sq.get("display_order", i),
            squad_type=sq.get("type", "product"),
            kpis_enabled=bool(sq.get("kpis_enabled", True)),
            budget_enabled=bool(sq.get("budget_enabled", False)),
            products=list(sq.get("products") or []),
            hardware=list(sq.get("hardware") or []),
        )
        squad = db.scalar(select(Squad).where(Squad.tribe_id == tribe.id, Squad.name == sq["name"]))
        if squad is None:
            squad = Squad(**fields)
            db.add(squad)
            db.flush()
            created["squads"] += 1
            log.info("Squad created: %s", squad.name)
        else:
            for k, v in fields.items():
                setattr(squad, k, v)
        squads_by_name[squad.name] = squad

    # --- Initiatives (tribe-level, optionally assigned to a squad) ---
    for i, it in enumerate(data.get("initiatives", []) or [], start=1):
        squad = squads_by_name.get(it.get("squad")) if it.get("squad") else None
        fields = dict(
            tribe_id=tribe.id, year=year, title=it["title"], description=it.get("description"),
            squad_id=squad.id if squad else None, owner=it.get("owner"),
            deadline=_to_datetime(it.get("deadline")), display_order=it.get("display_order", i),
        )
        init = db.scalar(select(Initiative).where(
            Initiative.tribe_id == tribe.id, Initiative.year == year, Initiative.title == it["title"]))
        if init is None:
            db.add(Initiative(**fields))
            created["initiatives"] += 1
            log.info("Initiative created: %s", it["title"])
        else:
            for k, v in fields.items():
                setattr(init, k, v)

    # --- OTDs (On-Time Delivery): the owner is the referenced squad's leader ---
    for i, o in enumerate(data.get("otds", []) or [], start=1):
        squad = squads_by_name.get(o.get("squad")) if o.get("squad") else None
        fields = dict(
            tribe_id=tribe.id, year=year, title=o["title"], description=o.get("description"),
            committed_date=_to_datetime(o.get("committed_date")),
            owner_user_id=squad.leader_user_id if squad else None,
            display_order=o.get("display_order", i),
        )
        otd = db.scalar(select(Otd).where(
            Otd.tribe_id == tribe.id, Otd.year == year, Otd.title == o["title"]))
        if otd is None:
            db.add(Otd(**fields))
            created["otds"] += 1
            log.info("OTD created: %s", o["title"])
        else:
            for k, v in fields.items():
                setattr(otd, k, v)

    db.commit()
    summary = {
        "tribe": tribe.name,
        "year": year,
        "squads": len(squads_by_name),
        "initiatives": len(data.get("initiatives") or []),
        "otds": len(data.get("otds") or []),
        "created": created,
    }
    log.info("Import terminé: %s", summary)
    return summary


# --------------------------------------------------------------------------- #
# Input reading (Excel or YAML) and Excel template generation
# --------------------------------------------------------------------------- #
def _yesno(v) -> bool:
    """Interpret an Excel cell as a boolean (oui/yes/true/1/x/vrai -> True)."""
    return str(v).strip().lower() in ("oui", "yes", "true", "1", "x", "vrai")


def _split(v) -> list[str]:
    """Split a comma-separated Excel cell into a clean list (empty -> [])."""
    if not v:
        return []
    return [p.strip() for p in str(v).split(",") if p.strip()]


def _read_xlsx(path: str) -> dict:
    """Parse the org workbook (sheets Tribu / Squads / Initiatives / OTD) into the
    same dict shape as the YAML input. Columns are read by POSITION, so the header
    text may be translated freely as long as the column order matches the template."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)

    def rows(sheet: str):
        if sheet not in wb.sheetnames:
            return []
        # Skip the header row; drop rows whose first cell is empty.
        return [r for r in wb[sheet].iter_rows(min_row=2, values_only=True) if r and r[0] not in (None, "")]

    tribe_rows = rows("Tribu")
    r = (list(tribe_rows[0]) + [None] * 5)[:5] if tribe_rows else [None] * 5
    data: dict = {
        "year": r[0],
        "tribe": {"name": r[1], "description": r[2], "leader": {"name": r[3], "email": r[4]}},
        "squads": [], "initiatives": [], "otds": [],
    }
    for row in rows("Squads"):
        name, typ, ln, le, products, hardware, kpis, budget = (list(row) + [None] * 8)[:8]
        data["squads"].append({
            "name": name, "type": (typ or "product"),
            "leader": {"name": ln, "email": le},
            "products": _split(products), "hardware": _split(hardware),
            "kpis_enabled": _yesno(kpis) if kpis is not None else True,
            "budget_enabled": _yesno(budget) if budget is not None else False,
        })
    for row in rows("Initiatives"):
        title, squad, owner, deadline, desc = (list(row) + [None] * 5)[:5]
        data["initiatives"].append({"title": title, "squad": squad, "owner": owner,
                                    "deadline": deadline, "description": desc})
    for row in rows("OTD"):
        title, squad, cdate, desc = (list(row) + [None] * 4)[:4]
        data["otds"].append({"title": title, "squad": squad, "committed_date": cdate, "description": desc})
    return data


def load_data(path: str) -> dict:
    """Read the org definition from an Excel (.xlsx) or YAML (.yaml/.yml) file."""
    if os.path.splitext(path)[1].lower() in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def read_upload(filename: str, content: bytes) -> dict:
    """Parse an uploaded file (in memory, no disk) into the org dict. Accepts an
    Excel (.xlsx/.xlsm) or YAML (.yaml/.yml) payload. Used by the admin import API
    so a filled file can be imported without rebuilding the image."""
    import io

    ext = os.path.splitext(filename or "")[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(io.BytesIO(content))
    if ext in (".yaml", ".yml"):
        return yaml.safe_load(content)
    raise ValueError("Format non supporte: fournissez un fichier .xlsx ou .yaml")


def build_template_workbook():
    """Build the empty-but-illustrated org workbook (4 sheets). Returned as an
    openpyxl ``Workbook`` so callers can save it to disk or stream it over HTTP."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E2761")

    def build(ws, title, headers, examples):
        ws.title = title
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font, cell.fill = head_font, head_fill
            ws.column_dimensions[cell.column_letter].width = max(16, len(h) + 4)
        for ri, ex in enumerate(examples, 2):
            for ci, v in enumerate(ex, 1):
                ws.cell(row=ri, column=ci, value=v)
        ws.freeze_panes = "A2"

    build(wb.active, "Tribu",
          ["Annee", "Nom de la tribu", "Description", "Tribe leader (nom)", "Tribe leader (email)"],
          [[2026, "Cloud Foundations", "Socle cloud souverain", "Nadia Khaldi", "nadia@exemple.com"]])
    build(wb.create_sheet(), "Squads",
          ["Nom", "Type (product/transverse)", "Squad leader (nom)", "Squad leader (email)",
           "Produits (separes par virgule)", "Materiel (separes par virgule)", "KPIs (oui/non)", "Budget (oui/non)"],
          [["Portal", "product", "Leo Martin", "leo@exemple.com", "Portal self-service", "", "oui", "non"],
           ["GCP / S3NS", "product", "Sara Dubois", "sara@exemple.com", "", "", "oui", "non"],
           ["Run & Operation", "transverse", "Paul Lemoine", "paul@exemple.com", "", "", "non", "non"]])
    build(wb.create_sheet(), "Initiatives",
          ["Titre", "Squad concernee", "Owner", "Echeance (AAAA-MM-JJ)", "Description"],
          [["Souverainete cloud & FinOps", "Portal", "Nadia Khaldi", "2026-12-31", "Reduire le cout du socle cloud."]])
    build(wb.create_sheet(), "OTD",
          ["Titre", "Squad concernee (owner = son leader)", "Date d'engagement (AAAA-MM-JJ)", "Description"],
          [["Data lake GCP en production", "GCP / S3NS", "2026-09-30", "Mise en production du data lake."]])
    return wb


def template_bytes() -> bytes:
    """Return the Excel template as raw bytes (for the download API)."""
    import io

    buf = io.BytesIO()
    build_template_workbook().save(buf)
    return buf.getvalue()


def write_template(path: str) -> None:
    """Write the Excel template (4 sheets) to ``path`` for the user to fill."""
    build_template_workbook().save(path)


def main(argv: list[str] | None = None) -> None:
    """CLI entry point.

    Usage:
        python -m app.import_org                 # import data/org.xlsx (or org.yaml)
        python -m app.import_org data/org.xlsx   # import a specific file
        python -m app.import_org --template      # (re)generate the Excel template
    """
    argv = sys.argv[1:] if argv is None else argv

    if argv and argv[0] == "--template":
        out = argv[1] if len(argv) > 1 else os.path.join(DATA_DIR, "org.template.xlsx")
        write_template(out)
        log.info("Modele Excel ecrit: %s", out)
        return

    path = argv[0] if argv else _default_path()
    if not os.path.isfile(path):
        log.error("Fichier introuvable: %s. Remplissez data/org.template.xlsx, "
                  "enregistrez-le sous data/org.xlsx, puis relancez.", path)
        sys.exit(1)
    data = load_data(path)
    if not data or "tribe" not in data or not (data.get("tribe") or {}).get("name"):
        log.error("Le fichier doit definir au moins une tribu (onglet 'Tribu' rempli).")
        sys.exit(1)
    db = SessionLocal()
    try:
        import_org(db, data)
    finally:
        db.close()


if __name__ == "__main__":
    main()
